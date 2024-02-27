# Internal imports
import csv
import argparse
import creds
import os

# External imports
import requests
# http://requests.readthedocs.org/en/latest/
import pyodbc
# https://github.com/mkleehammer/pyodbc/wiki
import openpyxl
# https://openpyxl.readthedocs.io/en/stable/


# -----------------------------
# Arguments
parser = argparse.ArgumentParser()

parser.add_argument("-t", "--tunnel",
                    dest="tunnel_needed",
                    action="store_true",
                    default=False,
                    help="Optional. Include to run the connection through a tunnel.")

parser.add_argument("-fy", "--ostifyfiles",
                    dest="process_osti_fy_files",
                    action="store_true",
                    default=False,
                    help="Optional. Process the OSTI FY xlsx files (upload these to /input).")

parser.add_argument("-g", "--googledrive",
                    dest="transfer_to_google_drive",
                    action="store_true",
                    default=False,
                    help="Optional. Uploads resulting files to the lbl tableau data google drive.")

parser.add_argument("-nt", "--newtoken",
                    dest="create_new_token",
                    action="store_true",
                    default=False,
                    help="TK TK not yet implemented: Optional. Creates a new google drive token.")

args = parser.parse_args()


# ========================================
def main():

    # Start the tunnel if needed
    if args.tunnel_needed:
        print("Opening SSH tunnel.")
        from sshtunnel import SSHTunnelForwarder

        server = SSHTunnelForwarder(
            creds.ssh_creds['host'],
            ssh_username=creds.ssh_creds['username'],
            ssh_password=creds.ssh_creds['password'],
            remote_bind_address=creds.ssh_creds['remote'],
            local_bind_address=creds.ssh_creds['local'])

        server.start()

    create_output_dir()

    # Connect to LBL's HR feed API and download 1 file
    get_lbl_hr_feed()

    # Connect to the reporting DB and download 3 files
    get_reporting_db_data()

    # Load the OSTI FY xlsx workbooks, export LBNL sheet as .csv
    if args.process_osti_fy_files:
        export_osti_fy_data()

    # Transfer output files to google drive
    if args.transfer_to_google_drive:
        transfer_to_google_drive()

    if args.tunnel_needed:
        server.stop()

    print("Program complete. Exiting.")


# ----------------------------------------
def create_output_dir():
    from datetime import datetime
    global output_dir

    output_dir = 'output/' + datetime.today().strftime('%Y-%m-%d-%H-%M-%S')
    os.mkdir(output_dir)

    print("Updated files will be output to: ", output_dir)


# ----------------------------------------
def output_json_to_csv(json_array, output_file_name):
    # Output file
    with open(output_dir + "/" + output_file_name, 'w') as csv_file:
        csv_writer = csv.writer(csv_file)

        # CSV header
        csv_writer.writerow(json_array[0].keys())

        # CSV rows
        for row in json_array:
            csv_writer.writerow(row.values())


# ----------------------------------------
def get_lbl_hr_feed():
    print("Connecting to LBL's API and downloading HR feed.")

    # Send the req, convert results to JSON
    result = requests.get(creds.lbl_api_endpoint, headers=creds.lbl_api_headers)
    result_json = result.json()

    # Output file
    output_json_to_csv(result_json, 'lbl_hr_feed.csv')


# ----------------------------------------
def get_reporting_db_data():
    print("Connecting to reporting DB...")
    sql_creds = creds.sql_creds_local_prod if args.tunnel_needed else creds.sql_creds_server_prod

    # Connect to Elements Reporting DB
    try:
        conn = pyodbc.connect(
            driver=sql_creds['driver'],
            server=(sql_creds['server'] + ',' + sql_creds['port']),
            database=sql_creds['database'],
            uid=sql_creds['user'],
            pwd=sql_creds['password'],
            trustservercertificate='yes')
    except:
        raise Exception("ERROR CONNECTING TO DATABASE.")
        exit(0)

    # Get the SQL query files, filter for anything non-SQL.
    query_files = os.listdir("sql_queries")
    query_files = [file for file in query_files if ".sql" in file]

    for query_file in query_files:
        print("Querying:", query_file)

        # Open a cursor, send the query.
        cursor = conn.cursor()
        sql_query = (open("sql_queries/" + query_file)).read()
        cursor.execute(sql_query)

        # pyodbc doesn't have a dict-cursor option, have to do it manually
        columns = [column[0] for column in cursor.description]
        json_array = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Output file
        output_file = query_file.replace(".sql", ".csv")
        output_json_to_csv(json_array, output_file)

        cursor.close()

    conn.close()


# ----------------------------------------
def export_osti_fy_data():
    print("Exporting/converting OSTI files")

    # Get the OSTI files, filter for any non-xlsx
    osti_files = os.listdir("input_osti_xlsx_files")
    osti_files = [file for file in osti_files if ".xlsx" in file]

    for osti_file in osti_files:
        wb = openpyxl.load_workbook("osti_files/" + osti_file)
        print(wb.sheetnames)

        if 'LBNL' not in wb.sheetnames:
            print("ERROR! There is no 'LBNL' sheet in the OSTI workbook.")
            exit(0)

        # Load the LBNL sheet, convert data to list
        lbnl_sheet = wb['LBNL']
        sheet_data = list(lbnl_sheet.rows)

        # Convert to list to an array of JSON
        sheet_json = []
        column_names = [column.value for column in sheet_data[0]]
        for row in sheet_data[1:]:
            values = [cell.value for cell in row]
            row_dict = {name: str(value) for name, value in zip(column_names, values)}
            sheet_json.append(row_dict)

        # Output file
        output_file = (osti_file.split(" - ")[0] + ".csv").lower()
        output_json_to_csv(sheet_json, output_file)


# ----------------------------------------
def transfer_to_google_drive():

    # import google drive packages
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaIoBaseDownload

    # If modifying these scopes, you need to create a new token.json.
    g_scopes = ["https://www.googleapis.com/auth/drive"]

    # This token may need to be refreshed at some point.
    # TK TK cross that bridge when needed
    if os.path.exists("token.json"):
        g_creds = Credentials.from_authorized_user_file("token.json", g_scopes)
    else:
        raise "No token.json found in this directory. Cannot connect to google drive."
        exit(1)

    try:
        # Create the google drive API service
        service = build("drive", "v3", credentials=g_creds)

        # Collect the files to load and reformat as dicts
        files_to_upload = [
            "lbl_hr_feed.csv",
            "Publication_Metadata.csv",
            "Unclaimed_Publication_Metadata.csv",
            "User_Pub_Relationships.csv"
        ]

        if args.process_osti_fy_files:
            files_to_upload.append(["fy2022.csv", "fy2023.csv", "fy2024.csv"])

        files_to_upload = [{'name': filename, 'id': None}
                           for filename in files_to_upload]

        # Get the file IDs in the g drive tableau data forlder
        # tableau_data_folder_id = '1HbkWZYiptaecVIXMUH3hv9zX4qrYUiWE'
        # Testing below
        tableau_data_folder_id = '13L_Au84OUuxReEVPw7Veh0WdSwF_3rwV'
        parent_folder_query = "'" + tableau_data_folder_id + "' in parents"
        results = (
            service.files().list(
                # pageSize=10,
                includeItemsFromAllDrives=True,
                includeTeamDriveItems=True,
                supportsAllDrives=True,
                supportsTeamDrives=True,
                q=parent_folder_query,
                fields="nextPageToken, files(id, name)",
            ).execute()
        )

        # Get the files and print a spot-check
        google_drive_files = results.get("files", [])
        print("G Drive tableau data folder files:", google_drive_files)

        # Assign extant IDs to files for upload
        for g_file in google_drive_files:
            for u_file in files_to_upload:
                if g_file['name'] == u_file['name']:
                    u_file['id'] = g_file['id']

        for upload_file in files_to_upload:
            upload_file_path = output_dir + "/" + upload_file['name']

            # No matching file in the folder, create a new file:
            if upload_file['id'] is None:
                metadata = {
                    'name': upload_file['name'],
                    'mimeType': 'text/csv',
                    'parents': [tableau_data_folder_id]}

                response = service.files().create(
                    body=metadata,
                    media_body=upload_file_path,
                ).execute()

                if response:
                    print("CREATED NEW FILE:", response)

            # The file already exists, update it with the new content.
            else:

                response = service.files().update(
                    fileId=upload_file['id'],
                    media_body=upload_file_path,
                ).execute()

                if response:
                    print("UPDATED FILE:", response)

    except HttpError as error:
        print(f"An error occurred: {error}")


# ----------------------------------------
if __name__ == '__main__':
    main()
